from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


def excel_write(begin_row,begin_line,value_list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    save_path = "../results.xlsx"
    booksheet = workbook.get_sheet_by_name("Sheet1")
    for i in range(len(value_list)):
        booksheet.cell(begin_row,begin_line + i).value = value_list[i]
        booksheet.cell(begin_row,begin_line + i).alignment = Alignment(horizontal='center', vertical='center')
    workbook.save(save_path)
